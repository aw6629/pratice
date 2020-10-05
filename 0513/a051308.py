import openpyxl
workbook=openpyxl.load_workbook(r'c:\tmp\aa.xlsx')
print(workbook.sheetnames)
active_worksheet=workbook.active
print(active_worksheet.title)
print(active_worksheet['A1'].value)
print(active_worksheet.cell(row=1,column=1).value)
print(active_worksheet.max_row)
print(active_worksheet.max_column)
